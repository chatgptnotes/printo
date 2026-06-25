"""
Bill of Quantities (BOQ) workbook generator.

Produces a regular, tender-style Excel workbook in the same layout as a
standard industry BOQ:

    Cover  ->  Preamble & Standards  ->  one Bill sheet per trade section
           ->  Summary of Bills

Styling matches a conventional tender BOQ: Arial throughout, navy section
bands, blue column headers, yellow (editable) Qty/Rate/Amount cells, light
grey zebra striping, thin cell borders. Each line's Amount is a live Excel
formula (= Qty x Rate); each Bill totals with =SUM; the Summary references
every Bill via a cross-sheet formula and rolls up to Sub-total, Provisional
Sums, Day Works, Contingency, Discount, Total excl. VAT, VAT and Grand Total.
No rates are invented here — the estimator types unit rates and the whole
workbook re-totals itself.
"""

from __future__ import annotations

import datetime
import io
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ── Palette / type (matches a conventional tender BOQ) ──────────────────────
FONT = "Arial"
NAVY = "1F3864"     # title / section bands / bill & grand totals (white bold)
BLUE = "2E75B6"     # column headers + sub-totals (white bold)
YELLOW = "FFF3CD"   # editable Qty / Rate / Amount cells
AMBER = "FFE0A3"    # unpriced Rate / Amount cells (flag for manual pricing)
ZEBRA = "F8F9FA"    # alternating item rows
CURRENCY_FMT = "#,##0.00"

_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Bill sheet columns (1-indexed) and widths, mirroring the reference sheet.
BILL_COLS = [
    ("Item", 7),
    ("Description", 58),
    ("Reference", 22),
    ("Unit", 8),
    ("Qty", 9),
    ("Rate\n(AED)", 13),
    ("Amount\n(AED)", 16),
    ("Origin / Brand", 32),
]
N_COLS = len(BILL_COLS)          # 8
LAST_COL = get_column_letter(N_COLS)   # "H"
ORIGIN_PLACEHOLDER = "— (contractor selection)"


# ── small helpers ───────────────────────────────────────────────────────────
def _val(v) -> str:
    if v is None:
        return "—"
    s = str(v).strip()
    return s or "—"


def _num(v):
    """Leading numeric value from a quantity string, else None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d[\d,]*\.?\d*", str(v))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def _boq_items(extracted: dict) -> list[dict]:
    return [b for b in (extracted.get("boq_items") or []) if isinstance(b, dict)]


def _sections(items: list[dict]) -> list[str]:
    out: list[str] = []
    for it in items:
        s = (it.get("section") or "General").strip() or "General"
        if s not in out:
            out.append(s)
    return out


def _sheet_title(prefix: str, name: str, used: set[str]) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", " ", name).strip()
    title = f"{prefix} {clean}".strip()[:31].rstrip()
    base, n = title, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[: 31 - len(suffix)].rstrip() + suffix
        n += 1
    used.add(title.lower())
    return title


def _project_line(extracted: dict, meta: dict) -> str:
    parts = [
        extracted.get("project_name") or meta.get("file_name") or "Project",
        extracted.get("project_location"),
        f"Drawing {extracted.get('drawing_number')}" if extracted.get("drawing_number") else None,
        "Authority / Client to confirm",
    ]
    return "   |   ".join(p for p in parts if p)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _f(size=10, bold=False, italic=False, color=None) -> Font:
    return Font(name=FONT, size=size, bold=bold, italic=italic, color=color)


def _put(ws, row, col, value=None, *, font=None, fill=None, align=None,
         border=True, numfmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or _f()
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = BORDER
    if numfmt:
        c.number_format = numfmt
    return c


def _band(ws, row, text, cols=N_COLS, *, fill=NAVY, size=12, height=27.8,
          align_left=True, color="FFFFFF"):
    """Full-width coloured band (title / section header / total)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _f(size=size, bold=True, color=color)
    c.fill = _fill(fill)
    c.alignment = Alignment(horizontal="left" if align_left else "center",
                            vertical="center", indent=1 if align_left else 0)
    ws.row_dimensions[row].height = height


def _print_setup(ws, header_rows, landscape=True):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = f"1:{header_rows}"
    ws.sheet_view.showGridLines = False


# ── Cover ────────────────────────────────────────────────────────────────────
def _build_cover(wb, extracted, meta, sections, *, approved, approved_by,
                 approved_at, generated_at):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 18, "C": 22, "D": 16, "E": 14, "F": 14,
                   "G": 14, "H": 22, "I": 3}.items():
        ws.column_dimensions[col].width = w

    # Title block — merged B2:H3 (navy) + B4:H4 subtitle (blue).
    ws.merge_cells("B2:H3")
    t = ws.cell(row=2, column=2, value="BILL OF QUANTITIES")
    t.font = _f(size=16, bold=True, color="FFFFFF")
    t.fill = _fill(NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (2, 3):
        for cc in range(2, 9):
            ws.cell(row=rr, column=cc).fill = _fill(NAVY)
        ws.row_dimensions[rr].height = 22
    subtitle = extracted.get("drawing_title") or "CONSTRUCTION WORKS — BILL OF QUANTITIES"
    ws.merge_cells("B4:H4")
    s = ws.cell(row=4, column=2, value=str(subtitle).upper())
    s.font = _f(size=10, bold=True, color="FFFFFF")
    s.fill = _fill(BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(2, 9):
        ws.cell(row=4, column=cc).fill = _fill(BLUE)
    ws.row_dimensions[4].height = 20

    def kv(row, label, value, height=16):
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = _f(bold=True)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        for cc in range(3, 9):
            ws.cell(row=row, column=cc).border = BORDER
        vc = ws.cell(row=row, column=3, value=_val(value))
        vc.font = _f()
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = height

    def header(row, text):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        c = ws.cell(row=row, column=2, value=text)
        c.font = _f(bold=True, color="FFFFFF")
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for cc in range(2, 9):
            ws.cell(row=row, column=cc).fill = _fill(NAVY)
        ws.row_dimensions[row].height = 18

    r = 6
    for label, key in (("Project", "project_name"), ("Owner / Client", "client_name"),
                       ("Location", "project_location"),
                       ("Consultant / Contractor", "contractor_name"),
                       ("Building Type", "building_type")):
        kv(r, label, extracted.get(key)); r += 1

    r += 1
    header(r, "TENDER DETAILS"); r += 1
    today = datetime.datetime.now().strftime("%d/%m/%Y")
    for label, value in (("Drawing No.", extracted.get("drawing_number")),
                         ("Drawing Date", extracted.get("date_of_issue") or today),
                         ("BOQ Date", today),
                         ("BOQ Status", "APPROVED" if approved else "TENDER — FOR PRICING"),
                         ("Currency", "AED (UAE Dirham)"),
                         ("VAT Rate", "5%")):
        kv(r, label, value); r += 1
    if approved and approved_by:
        when = str(approved_at or "")[:16].replace("T", " ")
        kv(r, "Approved By", f"{approved_by}{(' on ' + when) if when else ''}"); r += 1
    kv(r, "Generated", generated_at); r += 1

    r += 1
    header(r, "CONTRACTOR DETAILS"); r += 1
    for label in ("Contractor", "Address", "Phone", "Email", "TRN"):
        kv(r, label, ""); r += 1

    r += 1
    header(r, "LIST OF BILLS"); r += 1
    for col, text in ((2, "Bill"), (3, "Description"), (8, "Sheet")):
        if col == 3:
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        c = ws.cell(row=r, column=col, value=text)
        c.font = _f(bold=True, color="FFFFFF")
        c.fill = _fill(BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(2, 9):
        ws.cell(row=r, column=cc).border = BORDER
    r += 1
    for i, sec in enumerate(sections, 1):
        z = _fill(ZEBRA) if i % 2 == 0 else None
        bc = _put(ws, r, 2, i, align=Alignment(horizontal="center", vertical="center"), fill=z)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        _put(ws, r, 3, sec, align=Alignment(horizontal="left", vertical="center", wrap_text=True), fill=z)
        for cc in range(4, 8):
            _put(ws, r, cc, fill=z)
        _put(ws, r, 8, f"Bill {i}", align=Alignment(horizontal="center", vertical="center"), fill=z)
        r += 1

    _print_setup(ws, 0, landscape=False)


# ── Preamble & Standards (generic, professional) ────────────────────────────
def _build_preamble(wb):
    ws = wb.create_sheet("Preamble & Standards")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 96

    def band(row, text, height=22, size=12):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c = ws.cell(row=row, column=2, value=text)
        c.font = _f(size=size, bold=True, color="FFFFFF")
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = height

    def para(row, text):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c = ws.cell(row=row, column=2, value=text)
        c.font = _f(size=9.5)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 46

    def kv(row, a, b):
        ca = _put(ws, row, 2, a, font=_f(bold=True),
                  align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        cb = _put(ws, row, 3, b, font=_f(),
                  align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        ws.row_dimensions[row].height = 28

    band(2, "PREAMBLE & APPLICABLE STANDARDS", size=12, height=26)

    band(4, "1.  SCOPE OF WORKS", size=10, height=18)
    para(5, "Supply, installation, testing and commissioning of the complete works "
            "described in this Bill of Quantities, in accordance with the tender "
            "drawings, specifications, applicable standards and local authority "
            "requirements. The scope includes all materials, labour, plant, "
            "supervision, transport, builder's work in connection, testing, "
            "commissioning, as-built drawings and O&M manuals.")

    band(7, "2.  APPLICABLE STANDARDS", size=10, height=18)
    kv(8, "Standard", "Description")
    standards = [
        ("Project Specification", "Tender drawings & specifications take precedence where more onerous."),
        ("Local Building Codes", "National / municipal building, fire and electrical safety regulations."),
        ("IEC / BS / IS", "Relevant international and national product & installation standards."),
        ("Good Practice", "Manufacturer's instructions and good engineering practice throughout."),
    ]
    r = 9
    for a, b in standards:
        kv(r, a, b); r += 1

    r += 1
    band(r, "3.  PRICING NOTES", size=10, height=18); r += 1
    notes = [
        "(1) All rates include supply, delivery, off-loading, storage, installation, "
        "fixings, supports, identification labelling, testing, commissioning and handover.",
        "(2) Quantities are estimated from the tender drawings; the contractor shall "
        "verify on site before procurement.",
        "(3) Where a 'Sum' / 'Item' / 'Lot' unit is used, the rate covers the complete "
        "scope described including all incidentals.",
        "(4) Provisional sums are net; add overheads, profit and attendance separately "
        "where applicable.",
        "(5) All prices are exclusive of VAT; VAT is added at the foot of the Summary.",
        "(6) Rates remain firm and fixed for the contract duration unless agreed in writing.",
    ]
    for n in notes:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=n)
        c.font = _f(size=9.5)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 1

    _print_setup(ws, 0, landscape=False)


# ── one Bill sheet per trade section ────────────────────────────────────────
def _build_bill_sheet(wb, used, bill_no, section, items, extracted, meta):
    title = _sheet_title(f"Bill {bill_no} -", section, used)
    ws = wb.create_sheet(title)
    for i, (_, w) in enumerate(BILL_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _band(ws, 1, f"BILL No. {bill_no} — {section.upper()}", size=12, height=27.8)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    sub = ws.cell(row=2, column=1, value="Project: " + _project_line(extracted, meta))
    sub.font = _f(size=9, italic=True)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for cc in range(1, N_COLS + 1):
        ws.cell(row=2, column=cc).border = BORDER
    ws.row_dimensions[2].height = 14

    header_row = 3
    hdr_fill = _fill(BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (name, _) in enumerate(BILL_COLS, 1):
        _put(ws, header_row, i, name, font=_f(bold=True, color="FFFFFF"),
             fill=hdr_fill, align=hdr_align)
    ws.row_dimensions[header_row].height = 28

    reference = extracted.get("drawing_number") or "—"
    yellow = _fill(YELLOW)
    amber = _fill(AMBER)
    first = header_row + 1
    row = first
    for k, it in enumerate(items, 1):
        z = _fill(ZEBRA) if k % 2 == 0 else None
        qty = _num(it.get("quantity"))
        rate = _num(it.get("rate"))
        priced = rate is not None
        rate_fill = yellow if priced else amber   # amber = unpriced, flag for manual pricing
        _put(ws, row, 1, f"{bill_no}.{k}", font=_f(bold=True),
             align=Alignment(horizontal="left", vertical="top"), fill=z)
        _put(ws, row, 2, _val(it.get("description")), font=_f(),
             align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=z)
        _put(ws, row, 3, it.get("reference") or reference, font=_f(),
             align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=z)
        _put(ws, row, 4, _val(it.get("unit")), font=_f(),
             align=Alignment(horizontal="center", vertical="top"), fill=z)
        _put(ws, row, 5, qty if qty is not None else _val(it.get("quantity")),
             font=_f(), align=Alignment(horizontal="right", vertical="top"),
             fill=yellow, numfmt=(CURRENCY_FMT if qty is not None else None))
        _put(ws, row, 6, rate if priced else None, font=_f(),
             align=Alignment(horizontal="right", vertical="top"),
             fill=rate_fill, numfmt=CURRENCY_FMT)
        _put(ws, row, 7, f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})',
             font=_f(), align=Alignment(horizontal="right", vertical="top"),
             fill=rate_fill, numfmt=CURRENCY_FMT)
        _put(ws, row, 8, ("📋  " + (it.get("origin") or ORIGIN_PLACEHOLDER)),
             font=_f(), align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=z)
        row += 1

    last = row - 1
    if last < first:                       # empty-section guard
        ws.merge_cells(start_row=first, start_column=1, end_row=first, end_column=N_COLS)
        _put(ws, first, 1, "No line items in this section.", font=_f(italic=True),
             align=Alignment(horizontal="left", vertical="center", indent=1))
        for cc in range(1, N_COLS + 1):
            ws.cell(row=first, column=cc).border = BORDER
        last = first
        row = first + 1

    # Bill total (carried to Summary).
    total_row = row + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    lab = ws.cell(row=total_row, column=1,
                  value=f"BILL {bill_no} — TOTAL CARRIED TO SUMMARY  (excl. VAT)")
    lab.font = _f(bold=True, color="FFFFFF")
    lab.fill = _fill(NAVY)
    lab.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    tot = _put(ws, total_row, 7, f"=SUM(G{first}:G{last})",
               font=_f(bold=True, color="FFFFFF"), fill=_fill(NAVY),
               align=Alignment(horizontal="right", vertical="center"), numfmt=CURRENCY_FMT)
    _put(ws, total_row, 8, "", font=_f(bold=True, color="FFFFFF"), fill=_fill(NAVY))
    for cc in range(1, N_COLS + 1):
        ws.cell(row=total_row, column=cc).border = BORDER
    ws.row_dimensions[total_row].height = 22

    ws.auto_filter.ref = f"A{header_row}:{LAST_COL}{last}"
    ws.freeze_panes = f"A{first}"
    _print_setup(ws, header_row, landscape=False)   # portrait, fit-to-width
    return title, total_row


# ── Summary of Bills ────────────────────────────────────────────────────────
def _build_summary(wb, bill_refs, extracted, meta):
    ws = wb.create_sheet("Summary of Bills")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 8, "B": 56, "C": 14, "D": 22, "E": 35, "F": 5}.items():
        ws.column_dimensions[col].width = w

    _band(ws, 1, "BILL OF QUANTITIES — SUMMARY OF BILLS", cols=5, size=12,
          align_left=False)
    ws.merge_cells("A2:E2")
    sub = ws.cell(row=2, column=1, value=_project_line(extracted, meta))
    sub.font = _f(size=9, italic=True)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(1, 6):
        ws.cell(row=2, column=cc).border = BORDER
    ws.row_dimensions[2].height = 18

    hr = 4
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, text in enumerate(["Bill", "Description", "Sheet Ref",
                                "Bill Total (AED)", "Notes"], 1):
        _put(ws, hr, col, text, font=_f(bold=True, color="FFFFFF"),
             fill=_fill(BLUE), align=hdr_align)
    ws.row_dimensions[hr].height = 27.8

    row = hr + 1
    first = row
    for n, (sheet_title, total_row, section) in enumerate(bill_refs, 1):
        z = _fill(ZEBRA) if n % 2 == 0 else None
        _put(ws, row, 1, n, align=Alignment(horizontal="center", vertical="center"), fill=z)
        _put(ws, row, 2, section, align=Alignment(horizontal="left", vertical="center", wrap_text=True), fill=z)
        _put(ws, row, 3, f"Bill {n}", align=Alignment(horizontal="center", vertical="center"), fill=z)
        _put(ws, row, 4, f"='{sheet_title}'!G{total_row}",
             align=Alignment(horizontal="right", vertical="center"),
             fill=z, numfmt=CURRENCY_FMT)
        _put(ws, row, 5, "", align=Alignment(horizontal="left", vertical="center"), fill=z)
        row += 1
    last_bill = row - 1

    def fin(label, formula, *, navy=False, size=10):
        nonlocal row
        fill = _fill(NAVY if navy else BLUE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _f(size=size, bold=True, color="FFFFFF")
        lc.fill = fill
        lc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        vc = ws.cell(row=row, column=4, value=formula)
        vc.font = _f(size=size, bold=True, color="FFFFFF")
        vc.fill = fill
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.number_format = CURRENCY_FMT
        nc = ws.cell(row=row, column=5, value="")
        nc.font = _f(size=size, bold=True, color="FFFFFF")
        nc.fill = fill
        for cc in range(1, 6):
            ws.cell(row=row, column=cc).border = BORDER
        ws.row_dimensions[row].height = 18 if not navy else 24
        this = row
        row += 1
        return this

    sub_row = fin("SUB-TOTAL OF BILLS (excl. VAT)", f"=SUM(D{first}:D{last_bill})")
    prov = fin("Provisional Sums adjustment (if any)", 0)
    days = fin("Day Works Schedule (per separate schedule)", 0)
    cont = fin("Contingency (10%)", f"=D{sub_row}*0.1")
    disc = fin("Discount (if offered, enter as negative)", 0)
    excl = fin("TOTAL TENDER PRICE — Excluding VAT",
               f"=D{sub_row}+D{prov}+D{days}+D{cont}+D{disc}", navy=True, size=11)
    vat = fin("VAT (5%)", f"=D{excl}*0.05")
    fin("GRAND TOTAL TENDER PRICE — Including VAT", f"=D{excl}+D{vat}",
        navy=True, size=11)

    # Form of Tender.
    row += 1
    _band(ws, row, "FORM OF TENDER", cols=5, fill=NAVY, size=10, height=18); row += 1
    for line in (
        "We, the undersigned, having examined the drawings, specifications and this "
        "bill of quantities, offer to execute the works for the Grand Total stated above.",
        "Validity period: 90 days from tender submission.",
        "Construction period: as specified in the contract documents.",
        "Defects Liability Period: 12 months from substantial completion / handover.",
        "Signature & Stamp: ____________________________     Date: ______________",
    ):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=line)
        c.font = _f(size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        for cc in range(1, 6):
            ws.cell(row=row, column=cc).border = BORDER
        ws.row_dimensions[row].height = 26
        row += 1

    ws.freeze_panes = f"A{first}"
    _print_setup(ws, hr, landscape=False)


# ── public entry points ─────────────────────────────────────────────────────
def build_boq_workbook(report_data: dict, *, approved: bool = False,
                       approved_by: str | None = None,
                       approved_at: str | None = None) -> bytes:
    """Build a tender-style BOQ workbook for one drawing. Returns .xlsx bytes."""
    extracted = dict(report_data.get("extracted") or {})
    meta = report_data.get("drawing_meta") or {}
    items = _boq_items(extracted)
    sections = _sections(items) or ["General"]
    generated_at = datetime.datetime.now().strftime("%d %b %Y, %H:%M")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True   # force Excel to recompute every formula on open

    _build_cover(wb, extracted, meta, sections, approved=approved,
                 approved_by=approved_by, approved_at=approved_at,
                 generated_at=generated_at)
    _build_preamble(wb)

    used: set[str] = {"cover", "preamble & standards", "summary of bills"}
    bill_refs: list[tuple] = []
    for i, sec in enumerate(sections, 1):
        sec_items = [it for it in items
                     if ((it.get("section") or "General").strip() or "General") == sec]
        title, total_row = _build_bill_sheet(wb, used, i, sec, sec_items, extracted, meta)
        bill_refs.append((title, total_row, sec))

    _build_summary(wb, bill_refs, extracted, meta)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_project_workbook(drawings: list[dict]) -> bytes:
    """Combined workbook across drawings: an Overview sheet + one consolidated,
    filterable line-item sheet, in the same tender-BOQ style."""
    generated_at = datetime.datetime.now().strftime("%d %b %Y, %H:%M")
    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True   # force Excel to recompute every formula on open

    ov = wb.active
    ov.title = "Overview"
    ov.sheet_view.showGridLines = False
    for i, w in enumerate((8, 22, 40, 24, 14, 12), 1):
        ov.column_dimensions[get_column_letter(i)].width = w
    _band(ov, 1, "PROJECT BILL OF QUANTITIES — OVERVIEW", cols=6, size=14,
          align_left=False, height=24)
    ov.merge_cells("A2:F2")
    s = ov.cell(row=2, column=1,
                value=f"{len(drawings)} drawing(s)   |   generated {generated_at}")
    s.font = _f(size=9, italic=True)
    s.alignment = Alignment(horizontal="center", vertical="center")
    hr = 4
    for i, h in enumerate(["ID", "Drawing No.", "Title", "Project", "Floor", "BOQ Lines"], 1):
        _put(ov, hr, i, h, font=_f(bold=True, color="FFFFFF"), fill=_fill(BLUE),
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))

    cons = wb.create_sheet("Consolidated BOQ")
    cons.sheet_view.showGridLines = False
    cons_cols = [("Drawing", 16), ("Section", 22), ("Item", 7), ("Description", 50),
                 ("Reference", 18), ("Unit", 8), ("Qty", 9), ("Rate\n(AED)", 13),
                 ("Amount\n(AED)", 16), ("Origin / Brand", 28)]
    for i, (_, w) in enumerate(cons_cols, 1):
        cons.column_dimensions[get_column_letter(i)].width = w
    ncol = len(cons_cols)
    _band(cons, 1, "CONSOLIDATED BILL OF QUANTITIES — ALL DRAWINGS", cols=ncol,
          size=12, height=24)
    chr_ = 2
    for i, (name, _) in enumerate(cons_cols, 1):
        _put(cons, chr_, i, name, font=_f(bold=True, color="FFFFFF"), fill=_fill(BLUE),
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    cons.row_dimensions[chr_].height = 28

    yellow = _fill(YELLOW)
    amber = _fill(AMBER)
    crow, orow, z = chr_ + 1, hr + 1, 0
    for d in drawings:
        extracted = dict(d.get("extracted") or {})
        items = _boq_items(extracted)
        label = extracted.get("drawing_number") or d.get("file_name") or f"#{d.get('id')}"
        reference = extracted.get("drawing_number") or "—"
        oz = _fill(ZEBRA) if z % 2 else None
        _put(ov, orow, 1, d.get("id"), align=Alignment(horizontal="center", vertical="center"), fill=oz)
        _put(ov, orow, 2, _val(extracted.get("drawing_number")), align=Alignment(horizontal="left", vertical="center"), fill=oz)
        _put(ov, orow, 3, _val(extracted.get("drawing_title")), align=Alignment(horizontal="left", vertical="center", wrap_text=True), fill=oz)
        _put(ov, orow, 4, _val(extracted.get("project_name")), align=Alignment(horizontal="left", vertical="center", wrap_text=True), fill=oz)
        _put(ov, orow, 5, _val(d.get("floor_category")), align=Alignment(horizontal="center", vertical="center"), fill=oz)
        _put(ov, orow, 6, len(items), align=Alignment(horizontal="center", vertical="center"), fill=oz)
        orow += 1; z += 1
        for k, it in enumerate(items, 1):
            rz = _fill(ZEBRA) if k % 2 == 0 else None
            qty = _num(it.get("quantity"))
            rate = _num(it.get("rate"))
            priced = rate is not None
            rate_fill = yellow if priced else amber   # amber = unpriced, flag for manual pricing
            _put(cons, crow, 1, label, align=Alignment(horizontal="left", vertical="top"), fill=rz)
            _put(cons, crow, 2, it.get("section") or "General", align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=rz)
            _put(cons, crow, 3, k, align=Alignment(horizontal="center", vertical="top"), fill=rz)
            _put(cons, crow, 4, _val(it.get("description")), align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=rz)
            _put(cons, crow, 5, it.get("reference") or reference, align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=rz)
            _put(cons, crow, 6, _val(it.get("unit")), align=Alignment(horizontal="center", vertical="top"), fill=rz)
            _put(cons, crow, 7, qty if qty is not None else _val(it.get("quantity")),
                 align=Alignment(horizontal="right", vertical="top"), fill=yellow,
                 numfmt=(CURRENCY_FMT if qty is not None else None))
            _put(cons, crow, 8, rate if priced else None, align=Alignment(horizontal="right", vertical="top"), fill=rate_fill, numfmt=CURRENCY_FMT)
            _put(cons, crow, 9, f'=IF(OR(G{crow}="",H{crow}=""),"",G{crow}*H{crow})',
                 align=Alignment(horizontal="right", vertical="top"), fill=rate_fill, numfmt=CURRENCY_FMT)
            _put(cons, crow, 10, "📋  " + (it.get("origin") or ORIGIN_PLACEHOLDER),
                 align=Alignment(horizontal="left", vertical="top", wrap_text=True), fill=rz)
            crow += 1

    last_line = crow - 1
    if last_line >= chr_ + 1:
        gt = crow + 1
        ws_last = get_column_letter(ncol)
        cons.merge_cells(start_row=gt, start_column=1, end_row=gt, end_column=8)
        lab = cons.cell(row=gt, column=1, value="GRAND TOTAL — ALL DRAWINGS (excl. VAT)")
        lab.font = _f(bold=True, color="FFFFFF")
        lab.fill = _fill(NAVY)
        lab.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        g = _put(cons, gt, 9, f"=SUM(I{chr_ + 1}:I{last_line})",
                 font=_f(bold=True, color="FFFFFF"), fill=_fill(NAVY),
                 align=Alignment(horizontal="right", vertical="center"), numfmt=CURRENCY_FMT)
        _put(cons, gt, ncol, "", font=_f(bold=True, color="FFFFFF"), fill=_fill(NAVY))
        for cc in range(1, ncol + 1):
            cons.cell(row=gt, column=cc).border = BORDER
        cons.auto_filter.ref = f"A{chr_}:{ws_last}{last_line}"
        cons.freeze_panes = f"A{chr_ + 1}"
    _print_setup(cons, chr_, landscape=True)

    ov.freeze_panes = f"A{hr + 1}"
    _print_setup(ov, hr, landscape=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
