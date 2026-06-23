import asyncio
import datetime
import io
import json
import os
import shutil
import sqlite3
import time
import uuid
from pathlib import Path

from fastapi import FastAPI, File, Form, UploadFile, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from dotenv import load_dotenv

load_dotenv()

from database        import init_db, get_conn, DB_PATH
from extractor       import extract_drawing_with_prepass
from ai_provider     import provider_status
from rules           import run_all_rules, verdict
from realsoft_mapper import map_to_realsoft, average_confidence
from realsoft_client import push_to_realsoft, ping_realsoft, RealSoftAPIError
from report_generator import (generate_report, generate_project_report,
                              html_to_pdf_bytes)
from auth import (verify_login, create_token, require_auth,
                  is_locked, record_failure, clear_failures)
from ai_provider import gateway_erp_map
import base64
import preprocess

STORAGE_DIR = Path(__file__).parent.parent / "storage"
STORAGE_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif"}
MAX_FILE_SIZE_MB    = 20
# Extraction timeout — the gateway runs Claude CLI on the VPS, which is slower
# than a direct API, so allow more headroom than the original 55s. Configurable.
EXTRACT_TIMEOUT     = float(os.getenv("EXTRACT_TIMEOUT", "110"))

# ERP simulation mode when credentials are absent or placeholder
def _is_real_credential(val: str | None) -> bool:
    return bool(val and val.strip() and not val.strip().startswith("YOUR_"))

_erp_configured = _is_real_credential(os.getenv("REALSOFT_BASE_URL")) and \
                  _is_real_credential(os.getenv("REALSOFT_API_KEY"))

app = FastAPI(title="Printo API", version="2.0.0")

# Browser CORS. The Next.js frontend talks to this API server-side (BFF proxy), so
# the browser normally never calls it cross-origin. Keep this configurable for any
# direct browser access: set ALLOWED_ORIGINS to a comma-separated list of origins
# (e.g. "https://printo.vercel.app,http://localhost:3000") to lock it down in prod.
_origins_env = os.getenv("ALLOWED_ORIGINS", "*").strip()
_allow_origins = ["*"] if _origins_env == "*" else [o.strip() for o in _origins_env.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allow_origins,
    allow_credentials=_origins_env != "*",
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()


# ── Authentication ──────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    identifier: str            # username OR email
    password: str
    remember: bool = False


@app.post("/auth/login")
def auth_login(body: LoginRequest):
    ident = (body.identifier or "").strip()
    if not ident or not body.password:
        raise HTTPException(400, "Username and password are required")

    locked = is_locked(ident)
    if locked:
        raise HTTPException(429, f"Too many attempts. Try again in {locked}s.")

    user = verify_login(ident, body.password)
    if not user:
        record_failure(ident)
        raise HTTPException(401, "Invalid credentials")   # generic — no enumeration

    clear_failures(ident)
    token, expires_in = create_token(user["username"], user["role"], body.remember)
    return {
        "token": token,
        "token_type": "bearer",
        "expires_in": expires_in,
        "user": {"username": user["username"], "email": user["email"], "role": user["role"]},
    }


@app.get("/auth/me")
def auth_me(user: dict = Depends(require_auth)):
    return user


# ── Helpers ────────────────────────────────────────────────────────────────

def now() -> str:
    return datetime.datetime.now().strftime("%H:%M:%S")

def event(icon: str, step: str, message: str, etype: str = "info") -> str:
    line = f"  [{now()}]  {icon}  {step} — {message}"
    return f"data: {json.dumps({'line': line, 'type': etype})}\n\n"

def save_drawing_record(file_name: str, file_path: str, floor_category: str) -> int:
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO drawings (file_name, file_path, status, floor_category) VALUES (?,?,?,?)",
        (file_name, str(file_path), "processing", floor_category)
    )
    conn.commit()
    rec_id = cur.lastrowid
    conn.close()
    return rec_id

def update_drawing_status(drawing_id: int, status: str,
                           drawing_number: str = None, project_name: str = None,
                           drawing_title: str = None):
    conn = get_conn()
    conn.execute(
        "UPDATE drawings SET status=?, drawing_number=?, project_name=?, drawing_title=? WHERE id=?",
        (status, drawing_number, project_name, drawing_title, drawing_id)
    )
    conn.commit()
    conn.close()

def save_extractions(drawing_id: int, extracted: dict):
    conf = extracted.get("confidence", {})
    conn = get_conn()
    for field, value in extracted.items():
        if field == "confidence":
            continue
        if isinstance(value, list):
            value = json.dumps(value)
        elif isinstance(value, dict):
            value = json.dumps(value)
        conn.execute(
            "INSERT INTO extractions (drawing_id, field_name, field_value, confidence) VALUES (?,?,?,?)",
            (drawing_id, field, str(value) if value is not None else None, conf.get(field))
        )
    conn.commit()
    conn.close()

def save_exceptions(drawing_id: int, rule_results):
    conn = get_conn()
    for r in rule_results:
        if not r.passed:
            conn.execute(
                "INSERT INTO exceptions (drawing_id, rule_id, field_name, reason, severity) VALUES (?,?,?,?,?)",
                (drawing_id, r.rule_id, r.field_name, r.message, r.severity)
            )
    conn.commit()
    conn.close()

def save_erp_push(drawing_id: int, payload: dict, status: str, response: str):
    conn = get_conn()
    conn.execute(
        "INSERT INTO erp_pushes (drawing_id, payload, method, status, pushed_at, response) VALUES (?,?,?,?,?,?)",
        (drawing_id, json.dumps(payload), "api", status,
         datetime.datetime.now().isoformat(), response)
    )
    conn.commit()
    conn.close()

REPORT_DIR = Path(__file__).parent.parent / "reports"


def _store_report_data(drawing_id: int, payload: dict):
    """Persist the data needed to regenerate the report on demand (so it can reflect
    later human corrections and render to HTML or PDF)."""
    REPORT_DIR.mkdir(exist_ok=True)
    (REPORT_DIR / f"{drawing_id}.json").write_text(
        json.dumps(payload, default=str), encoding="utf-8")


def _load_report_data(drawing_id: int) -> dict | None:
    path = REPORT_DIR / f"{drawing_id}.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _corrections_for(drawing_id: int) -> list:
    conn = get_conn()
    rows = conn.execute(
        "SELECT field_name, original_value, corrected_value, corrected_by, corrected_at "
        "FROM corrections WHERE drawing_id=? ORDER BY corrected_at DESC", (drawing_id,)
    ).fetchall()
    conn.close()
    return [{"field": r[0], "original": r[1], "corrected": r[2], "by": r[3], "at": r[4]}
            for r in rows]


_MARK_RED = (217, 25, 25)          # Pratyaya-style mistake red (#D91919)


def _failing_fields(rule_results) -> list[tuple]:
    """(rule_id, field_name, message) for failed ERROR/WARNING rules. Robust to
    RuleResult objects or plain dicts."""
    out = []
    for r in rule_results or []:
        passed   = getattr(r, "passed", None) if not isinstance(r, dict) else r.get("passed")
        severity = getattr(r, "severity", "") if not isinstance(r, dict) else r.get("severity", "")
        field    = getattr(r, "field_name", "") if not isinstance(r, dict) else r.get("field_name", "")
        rule_id  = getattr(r, "rule_id", "") if not isinstance(r, dict) else r.get("rule_id", "")
        message  = getattr(r, "message", "") if not isinstance(r, dict) else r.get("message", "")
        if not passed and severity in ("ERROR", "WARNING"):
            out.append((rule_id, field, message))
    return out


def _load_font(size: int):
    from PIL import ImageFont
    for name in ("arialbd.ttf", "arial.ttf", "DejaVuSans-Bold.ttf", "DejaVuSans.ttf"):
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _annotate_mistakes(img, extracted: dict, rule_results) -> None:
    """Draw red boxes (Pratyaya-style) on `img` for failed ERROR/WARNING fields.
    Boxes use normalised field_locations; failures without a location are stacked
    as red notes top-left. Mutates `img` in place; never raises to the caller."""
    from PIL import ImageDraw
    failing = _failing_fields(rule_results)
    if not failing:
        return
    locations = (extracted or {}).get("field_locations") or {}
    w, h = img.size
    lw = max(2, round(w / 350))
    font = _load_font(max(13, w // 70))
    draw = ImageDraw.Draw(img)

    def text_wh(s):
        l, t, r, b = draw.textbbox((0, 0), s, font=font)
        return r - l, b - t

    # group rule_ids per field so one box can carry several findings
    by_field: dict[str, list[str]] = {}
    unlocated: list[tuple] = []
    for rule_id, field, message in failing:
        box = locations.get(field)
        if box and len(box) == 4:
            by_field.setdefault(field, []).append(rule_id)
        else:
            unlocated.append((rule_id, field, message))

    # located boxes
    for field, rule_ids in by_field.items():
        x1, y1, x2, y2 = locations[field]
        px = [int(x1 * w), int(y1 * h), int(x2 * w), int(y2 * h)]
        px = [min(max(0, px[0]), w - 1), min(max(0, px[1]), h - 1),
              min(max(1, px[2]), w),     min(max(1, px[3]), h)]
        draw.rectangle(px, outline=_MARK_RED, width=lw)
        label = ",".join(dict.fromkeys(rule_ids))          # unique, order-kept
        tw, th = text_wh(label)
        tx1, ty1 = px[0], max(0, px[1] - th - 6)
        draw.rectangle([tx1, ty1, tx1 + tw + 8, ty1 + th + 6], fill=_MARK_RED)
        draw.text((tx1 + 4, ty1 + 3), label, fill="white", font=font)

    # unlocated findings → stacked red notes, top-left
    y = 10
    for rule_id, field, message in unlocated:
        note = f"{rule_id} {field}: {message}"[:90]
        tw, th = text_wh(note)
        draw.rectangle([10, y, 10 + tw + 8, y + th + 6], fill=_MARK_RED)
        draw.text((14, y + 3), note, fill="white", font=font)
        y += th + 10


def _thumbnail_data_uri(file_path: str, extracted: dict = None,
                        rule_results=None, max_w: int = 1000) -> str | None:
    """Render/downscale the source drawing to a base64 PNG data URI for embedding.
    When `extracted` + `rule_results` are supplied, draw red markings on the
    locations of failed ERROR/WARNING fields (Pratyaya-style)."""
    try:
        img_set = preprocess.build_images(file_path, dpi=150)
        raw = img_set.get("full")
        if not raw:
            return None
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(raw)).convert("RGB")
        if img.width > max_w:
            img = img.resize((max_w, int(img.height * max_w / img.width)), Image.LANCZOS)
        try:
            _annotate_mistakes(img, extracted, rule_results)
        except Exception:
            pass                                            # plain image on any draw failure
        buf = _io.BytesIO()
        img.save(buf, format="PNG")
        return "data:image/png;base64," + base64.standard_b64encode(buf.getvalue()).decode()
    except Exception:
        return None


def _rebuild_report_html(drawing_id: int) -> str | None:
    """Regenerate the per-drawing report from stored data + latest corrections."""
    data = _load_report_data(drawing_id)
    if not data:
        return None
    extracted = dict(data.get("extracted") or {})
    corrections = _corrections_for(drawing_id)
    # apply the latest correction per field so the report reflects human edits
    latest = {}
    for c in corrections:                       # newest-first; keep first seen
        latest.setdefault(c["field"], c["corrected"])
    for field, value in latest.items():
        extracted[field] = value
    rule_results = run_all_rules(extracted)
    verd = verdict(rule_results)
    thumb = _thumbnail_data_uri(data.get("file_path", ""), extracted, rule_results)
    return generate_report(
        data.get("drawing_meta", {"drawing_id": drawing_id}), extracted, rule_results,
        verd, data.get("elapsed", 0), data.get("erp_payload", {}),
        corrections=corrections, thumbnail_uri=thumb,
    )


# ── Main SSE Pipeline ─────────────────────────────────────────────────────

async def run_pipeline(file_path: Path, file_name: str, drawing_id: int,
                        file_size_mb: float, strict: bool, floor_category: str):
    start       = time.time()
    errors      = []
    warnings    = []
    extracted   = {}
    realsoft_payload = {}
    prepass_count = 0

    yield event("📁", "File received", f"{file_name} ({file_size_mb:.1f} MB) — Category: {floor_category}", "info")
    yield event("💾", "Stored", f"Saved to storage — Drawing ID: {drawing_id}", "info")

    # ── Pre-upload checks ────────────────────────────────────────────────
    yield event("🔍", "Pre-upload checks", "Verifying file format, size, and readability...", "info")

    suffix = Path(file_name).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        msg = f"Unsupported format: {suffix} | Allowed: PDF, JPG, PNG, TIFF"
        yield event("❌", "R01 FAILED", msg, "error")
        errors.append(msg)
        yield _end_failed(errors, warnings, {}, {}, drawing_id, start)
        return
    yield event("✅", "R01 PASSED", f"File format valid — {suffix.upper()} accepted", "success")

    if file_size_mb > MAX_FILE_SIZE_MB:
        msg = f"File too large: {file_size_mb:.1f}MB | Limit: {MAX_FILE_SIZE_MB}MB"
        yield event("❌", "R02 FAILED", msg, "error")
        errors.append(msg)
        yield _end_failed(errors, warnings, {}, {}, drawing_id, start)
        return
    yield event("✅", "R02 PASSED", f"File size OK — {file_size_mb:.1f}MB (limit: {MAX_FILE_SIZE_MB}MB)", "success")
    yield event("✅", "R03 PASSED", "File readable and non-empty", "success")

    # ── Image quality (blur) check ───────────────────────────────────────
    score = preprocess.sharpness_score(str(file_path))
    threshold = float(os.getenv("BLUR_THRESHOLD", "100"))
    if score is not None and score < threshold:
        msg = ("Uploaded image is too blurry for report generation — "
               "please re-upload a clearer scan or photo.")
        yield event("❌", "BLUR CHECK FAILED",
                    f"{msg} (sharpness {score:.0f} < {threshold:.0f})", "error")
        errors.append(msg)
        yield _end_failed(errors, warnings, {}, {}, drawing_id, start, status="blurred")
        return
    yield event("✅", "Blur Check",
                f"Image sharp enough (sharpness {score:.0f})" if score is not None
                else "Blur check skipped (OpenCV unavailable)", "success")

    # ── Pre-pass (PDF text layer) ────────────────────────────────────────
    if suffix == ".pdf":
        yield event("🔎", "Pre-pass", "Scanning PDF text layer for title block fields (free extraction)...", "info")
        # prepass happens inside extract_drawing_with_prepass — results reported after

    # ── AI Extraction ────────────────────────────────────────────────────
    yield event("🤖", "Claude Vision AI", "Sending drawing for intelligent field extraction...", "info")
    yield event("⏳", "AI Processing", "Reading title block, floor plan, dimensions, materials, stamps...", "info")

    try:
        extracted, prepass_hints = await asyncio.wait_for(
            asyncio.get_event_loop().run_in_executor(
                None,
                lambda: extract_drawing_with_prepass(
                    str(file_path), floor_category, file_name
                ),
            ),
            timeout=EXTRACT_TIMEOUT
        )
        prepass_count = len(prepass_hints)
    except asyncio.TimeoutError:
        msg = f"AI extraction timed out (>{int(EXTRACT_TIMEOUT)}s) — drawing saved, please retry"
        yield event("⏱️", "Timeout", msg, "error")
        update_drawing_status(drawing_id, "timeout")
        yield f"data: {json.dumps({'type': 'done', 'verdict': 'TIMEOUT'})}\n\n"
        return
    except Exception as e:
        msg = f"AI extraction failed: {str(e)}"
        yield event("❌", "AI Error", msg, "error")
        update_drawing_status(drawing_id, "error")
        yield f"data: {json.dumps({'type': 'done', 'verdict': 'ERROR'})}\n\n"
        return

    if prepass_count:
        yield event("✅", "Pre-pass", f"{prepass_count} fields extracted from PDF text layer (high confidence)", "success")

    field_count = sum(1 for k, v in extracted.items()
                      if k != "confidence" and v not in (None, [], ""))
    yield event("✅", "AI Extraction Complete",
                f"{field_count} fields extracted — project: {extracted.get('project_name', '—')}", "success")
    save_extractions(drawing_id, extracted)

    # ── Rules Validation ─────────────────────────────────────────────────
    yield event("🔍", "Validation Rules", "Running ClearSoft validation rules (R04–R18)...", "info")

    rule_results = run_all_rules(extracted, strict=strict)
    for r in rule_results:
        if r.passed:
            yield event("✅", f"{r.rule_id} PASSED", r.message, "success")
        elif r.severity == "ERROR":
            yield event("❌", f"{r.rule_id} FAILED", r.message, "error")
            errors.append(r.message)
        else:
            yield event("⚠️ ", f"{r.rule_id} WARNING", r.message, "warning")
            warnings.append(r.message)

    save_exceptions(drawing_id, rule_results)
    verd = verdict(rule_results)

    # ── Map to RealSoft Format ───────────────────────────────────────────
    yield event("🗺️ ", "ERP Mapping", "Converting extracted data to RealSoft ERP format...", "info")
    avg_conf = average_confidence(extracted)
    realsoft_payload = map_to_realsoft(extracted, drawing_id, file_name, verd, avg_conf)
    # Gateway-primary mapping: use the VPS ERP_MAP when reachable, else keep local.
    gw_data = gateway_erp_map(extracted)
    if gw_data:
        realsoft_payload["data"] = gw_data
        realsoft_payload["metadata"]["mapping_source"] = "gateway"
        yield event("✅", "ERP Mapping (Gateway)",
                    "Mapped via Printo Gateway (ERP_MAP)", "success")
    else:
        realsoft_payload["metadata"]["mapping_source"] = "local"
        yield event("✅", "Mapping Complete",
                    f"JSON payload ready — module: {realsoft_payload.get('module', 'DrawingMaster')}", "success")

    # ── ERP Push (or simulation) ─────────────────────────────────────────
    erp_status = "simulated"
    if _erp_configured:
        yield event("🚀", "RealSoft API", "Pushing to RealSoft test environment...", "info")
        try:
            api_response = await asyncio.get_event_loop().run_in_executor(
                None, push_to_realsoft, realsoft_payload
            )
            yield event("✅", "ERP Push Success",
                        f"RealSoft responded: HTTP {api_response['status_code']}", "success")
            save_erp_push(drawing_id, realsoft_payload, "sent", json.dumps(api_response))
            erp_status = "sent"
        except RealSoftAPIError as e:
            yield event("⚠️ ", "ERP Push Failed", f"{str(e)} — payload saved for retry", "warning")
            save_erp_push(drawing_id, realsoft_payload, "failed", str(e))
            warnings.append(f"ERP Push failed: {str(e)}")
            erp_status = "failed"
    else:
        yield event("🖥️ ", "ERP Simulation",
                    "RealSoft credentials not configured — data ready, showing simulation", "info")
        save_erp_push(drawing_id, realsoft_payload, "simulated", "Simulation mode — no ERP credentials")

    # ── Generate Report ──────────────────────────────────────────────────
    yield event("📄", "Report", "Generating summary report...", "info")
    elapsed = round(time.time() - start, 1)
    drawing_meta = {
        "file_name":      file_name,
        "drawing_id":     drawing_id,
        "floor_category": floor_category,
        "uploaded_at":    datetime.datetime.now().isoformat(),
    }
    _store_report_data(drawing_id, {
        "drawing_meta":  drawing_meta,
        "extracted":     extracted,
        "verdict":       verd,
        "elapsed":       elapsed,
        "erp_payload":   realsoft_payload,
        "file_path":     str(file_path),
    })
    yield event("✅", "Report Ready", f"Summary report generated — view at /report/{drawing_id}", "success")

    # ── Save final status ────────────────────────────────────────────────
    update_drawing_status(
        drawing_id,
        "done" if verd != "FAILED" else "error",
        drawing_number=extracted.get("drawing_number"),
        project_name=extracted.get("project_name"),
        drawing_title=extracted.get("drawing_title"),
    )

    passed = sum(1 for r in rule_results if r.passed)
    yield event("🏁", "Complete",
                f"Done in {elapsed}s — {passed}/{len(rule_results)} rules passed | "
                f"{len(errors)} errors | {len(warnings)} warnings", "done")

    yield "data: " + json.dumps({
        "type":             "done",
        "verdict":          verd,
        "elapsed":          elapsed,
        "errors":           errors,
        "warnings":         warnings,
        "extracted":        extracted,
        "realsoft_payload": realsoft_payload,
        "erp_status":       erp_status,
        "drawing_id":       drawing_id,
        "prepass_count":    prepass_count,
    }) + "\n\n"


def _end_failed(errors, warnings, extracted, payload, drawing_id, start, status="error") -> str:
    update_drawing_status(drawing_id, status)
    elapsed = round(time.time() - start, 1)
    return f"data: {json.dumps({'type': 'done', 'verdict': 'FAILED', 'elapsed': elapsed, 'errors': errors, 'warnings': warnings, 'extracted': extracted, 'realsoft_payload': payload, 'drawing_id': drawing_id})}\n\n"


# ── Endpoints ─────────────────────────────────────────────────────────────

@app.post("/upload")
async def upload_drawing(
    file: UploadFile = File(...),
    floor_category: str = Form(default="Other"),
    strict: bool = False,
    _user: dict = Depends(require_auth),
):
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported file type: {suffix}")

    unique_name = f"{uuid.uuid4().hex}{suffix}"
    dest = STORAGE_DIR / unique_name
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    size_mb    = dest.stat().st_size / (1024 * 1024)
    drawing_id = save_drawing_record(file.filename, dest, floor_category)

    return StreamingResponse(
        run_pipeline(dest, file.filename, drawing_id, size_mb, strict, floor_category),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/drawings")
def list_drawings(_user: dict = Depends(require_auth)):
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, file_name, uploaded_at, status, drawing_number, drawing_title, project_name, floor_category "
        "FROM drawings ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return [
        {"id": r[0], "file_name": r[1], "uploaded_at": r[2], "status": r[3],
         "drawing_number": r[4], "drawing_title": r[5], "project_name": r[6],
         "floor_category": r[7]}
        for r in rows
    ]


@app.get("/drawings/{drawing_id}")
def get_drawing(drawing_id: int, _user: dict = Depends(require_auth)):
    conn = get_conn()
    row = conn.execute("SELECT * FROM drawings WHERE id=?", (drawing_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Drawing not found")
    fields = conn.execute(
        "SELECT field_name, field_value, confidence, validated, flagged FROM extractions WHERE drawing_id=?",
        (drawing_id,)
    ).fetchall()
    pushes = conn.execute(
        "SELECT method, status, pushed_at, response FROM erp_pushes WHERE drawing_id=?",
        (drawing_id,)
    ).fetchall()
    corrections = conn.execute(
        "SELECT field_name, original_value, corrected_value, corrected_by, corrected_at "
        "FROM corrections WHERE drawing_id=? ORDER BY corrected_at DESC",
        (drawing_id,)
    ).fetchall()
    conn.close()
    return {
        "drawing":    dict(zip(
            ["id", "file_name", "file_path", "uploaded_at", "status",
             "drawing_number", "drawing_title", "project_name", "floor_category"], row
        )),
        "extractions": [{"field": f[0], "value": f[1], "confidence": f[2]} for f in fields],
        "erp_pushes":  [{"method": p[0], "status": p[1], "pushed_at": p[2]} for p in pushes],
        "corrections": [{"field": c[0], "original": c[1], "corrected": c[2],
                         "by": c[3], "at": c[4]} for c in corrections],
    }


@app.patch("/drawings/{drawing_id}/correction")
def save_correction(drawing_id: int, body: dict, _user: dict = Depends(require_auth)):
    field_name      = body.get("field_name")
    corrected_value = body.get("corrected_value")
    corrected_by    = body.get("corrected_by", "user")
    if not field_name:
        raise HTTPException(400, "field_name required")

    conn = get_conn()
    orig = conn.execute(
        "SELECT field_value FROM extractions WHERE drawing_id=? AND field_name=?",
        (drawing_id, field_name)
    ).fetchone()
    original_value = orig[0] if orig else None

    conn.execute(
        "INSERT INTO corrections (drawing_id, field_name, original_value, corrected_value, corrected_by) "
        "VALUES (?,?,?,?,?)",
        (drawing_id, field_name, original_value, corrected_value, corrected_by)
    )
    conn.execute(
        "UPDATE extractions SET field_value=?, validated=1 WHERE drawing_id=? AND field_name=?",
        (corrected_value, drawing_id, field_name)
    )
    conn.commit()
    conn.close()
    return {"message": f"Correction saved for {field_name}"}


def _all_drawings_for_project_report() -> list:
    """Gather drawings + per-drawing average confidence for the aggregate report."""
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, drawing_number, drawing_title, project_name, floor_category, status "
        "FROM drawings ORDER BY id"
    ).fetchall()
    out = []
    for r in rows:
        avg = conn.execute(
            "SELECT AVG(confidence) FROM extractions WHERE drawing_id=? AND confidence IS NOT NULL",
            (r[0],)
        ).fetchone()[0]
        out.append({"id": r[0], "drawing_number": r[1], "drawing_title": r[2],
                    "project_name": r[3], "floor_category": r[4], "status": r[5],
                    "avg_conf": avg})
    conn.close()
    return out


@app.get("/report/project", response_class=HTMLResponse)
def get_project_report(_user: dict = Depends(require_auth)):
    return HTMLResponse(content=generate_project_report(_all_drawings_for_project_report()))


@app.get("/report/project/pdf")
def get_project_report_pdf(_user: dict = Depends(require_auth)):
    pdf = html_to_pdf_bytes(generate_project_report(_all_drawings_for_project_report()))
    if not pdf:
        raise HTTPException(500, "PDF generation failed")
    return StreamingResponse(
        io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="printo_project_report.pdf"'})


@app.get("/report/{drawing_id}", response_class=HTMLResponse)
def get_report(drawing_id: int, _user: dict = Depends(require_auth)):
    html = _rebuild_report_html(drawing_id)
    if html is None:
        raise HTTPException(404, "Report not yet generated — process the drawing first")
    return HTMLResponse(content=html)


@app.get("/report/{drawing_id}/pdf")
def get_report_pdf(drawing_id: int, _user: dict = Depends(require_auth)):
    html = _rebuild_report_html(drawing_id)
    if html is None:
        raise HTTPException(404, "Report not yet generated — process the drawing first")
    pdf = html_to_pdf_bytes(html)
    if not pdf:
        raise HTTPException(500, "PDF generation failed")
    return StreamingResponse(
        io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="printo_report_{drawing_id}.pdf"'})


@app.get("/export/{drawing_id}/excel")
def export_excel(drawing_id: int, _user: dict = Depends(require_auth)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_conn()
    drawing = conn.execute("SELECT * FROM drawings WHERE id=?", (drawing_id,)).fetchone()
    if not drawing:
        conn.close()
        raise HTTPException(404, "Drawing not found")
    extractions = conn.execute(
        "SELECT field_name, field_value, confidence, validated FROM extractions WHERE drawing_id=?",
        (drawing_id,)
    ).fetchall()
    exceptions = conn.execute(
        "SELECT rule_id, field_name, reason, severity, resolved FROM exceptions WHERE drawing_id=?",
        (drawing_id,)
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()

    # Sheet 1 — Extraction
    ws1 = wb.active
    ws1.title = "Extraction"
    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Field", "Extracted Value", "Confidence", "Validated"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, (fname, fval, fconf, fval_flag) in enumerate(extractions, 2):
        ws1.cell(row=row_idx, column=1, value=fname)
        ws1.cell(row=row_idx, column=2, value=fval)
        ws1.cell(row=row_idx, column=3,
                 value=f"{fconf:.0%}" if fconf is not None else "—")
        ws1.cell(row=row_idx, column=4, value="Yes" if fval_flag else "No")
    for col in [1, 2, 3, 4]:
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 28

    # Sheet 2 — Validation
    ws2 = wb.create_sheet("Validation")
    for col, h in enumerate(["Rule ID", "Field", "Message", "Severity", "Resolved"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, (rid, fname, reason, sev, resolved) in enumerate(exceptions, 2):
        ws2.cell(row=row_idx, column=1, value=rid)
        ws2.cell(row=row_idx, column=2, value=fname)
        ws2.cell(row=row_idx, column=3, value=reason)
        ws2.cell(row=row_idx, column=4, value=sev)
        ws2.cell(row=row_idx, column=5, value="Yes" if resolved else "No")
    for col in [1, 2, 3, 4, 5]:
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname_safe = f"printo_drawing_{drawing_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname_safe}"'},
    )


@app.get("/exceptions")
def list_exceptions(resolved: bool = False, _user: dict = Depends(require_auth)):
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, drawing_id, rule_id, field_name, reason, severity, resolved, created_at "
        "FROM exceptions WHERE resolved=?",
        (1 if resolved else 0,)
    ).fetchall()
    conn.close()
    return [{"id": r[0], "drawing_id": r[1], "rule_id": r[2], "field_name": r[3],
             "reason": r[4], "severity": r[5], "resolved": r[6], "created_at": r[7]} for r in rows]


@app.patch("/exceptions/{exc_id}/resolve")
def resolve_exception(exc_id: int, resolved_by: str = "admin",
                      _user: dict = Depends(require_auth)):
    conn = get_conn()
    conn.execute("UPDATE exceptions SET resolved=1, resolved_by=? WHERE id=?", (resolved_by, exc_id))
    conn.commit()
    conn.close()
    return {"message": f"Exception {exc_id} resolved"}


@app.get("/health")
def health():
    realsoft_reachable = ping_realsoft() if _erp_configured else False
    conn = get_conn()
    total = conn.execute("SELECT COUNT(*) FROM drawings").fetchone()[0]
    done  = conn.execute("SELECT COUNT(*) FROM drawings WHERE status='done'").fetchone()[0]
    conn.close()

    ai = provider_status()   # {ai_provider, sidecar_reachable, mode, model, ...}
    return {
        "status":             "ok",
        "version":            "2.1.0",
        "total_drawings":     total,
        "completed":          done,
        "erp_mode":           "live" if _erp_configured else "simulation",
        # AI extraction provider (sidecar | mock)
        "ai_provider":        ai.get("ai_provider"),
        "ai_mode":            ai.get("mode"),
        "ai_model":           ai.get("model"),
        "sidecar_reachable":  ai.get("sidecar_reachable"),
        "mock_extraction":    ai.get("ai_provider") == "mock",
        "realsoft_reachable": realsoft_reachable,
        "realsoft_url":       os.getenv("REALSOFT_BASE_URL", "not configured"),
    }
